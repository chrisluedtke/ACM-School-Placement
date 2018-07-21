import datetime
import csv
import os

from .cleaning import *
from .commutes import *
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import render, reverse
from django.utils import timezone
from .forms import RunParametersForm
from .models import RunParameters
from openpyxl import load_workbook

plcmt_path = os.path.join(settings.MEDIA_ROOT, "documents/outputs/Output_Placements.csv")
trace_path = os.path.join(settings.MEDIA_ROOT, "documents/outputs/Output_Trace.csv")
commute_path = os.path.join(settings.MEDIA_ROOT, "documents/outputs/Output_Commute_Reference.csv")
error_path = os.path.join(settings.MEDIA_ROOT, "documents/outputs/errors.txt")
rm8_path = os.path.join(settings.MEDIA_ROOT, "documents/outputs/Invalid Roommate and Prior Relationship Names.csv")

params_path = os.path.join(settings.MEDIA_ROOT, "documents/params.csv")
survey_data_path = os.path.join(settings.MEDIA_ROOT, "documents/ACM_Placement_Survey_Data.csv")
school_data_path = os.path.join(settings.MEDIA_ROOT, "documents/ACM_Placement_School_Data.xlsx")

def reset_files():
    for file in [plcmt_path, trace_path, error_path, rm8_path, commute_path, school_data_path, survey_data_path, params_path]:
        if os.path.exists(file):
            os.remove(file)

# Create your views here.
def welcome(request):
    reset_files()
    return render(request, 'procedure/welcome.html')

def step1(request):
    reset_files()

    # download school data template
    if request.method == 'POST' and 'download' in request.POST:
        file_path = os.path.join(settings.MEDIA_ROOT, "documents/template/ACM_Placement_School_Data.xlsx")
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
            return response
    # upload school data template
    if request.method == 'POST' and 'upload' in request.POST and 'myfile' in request.FILES and '.xlsx' in request.FILES['myfile'].name:
        filename = FileSystemStorage().save(school_data_path, request.FILES['myfile'])
        return HttpResponseRedirect(reverse('step2'))
    return render(request, 'procedure/step1.html', {})

def step2(request):
    if os.path.exists(survey_data_path):
        os.remove(survey_data_path)
    # Upload ACM data
    if request.method == 'POST' and 'upload' in request.POST and 'myfile' in request.FILES and '.csv' in request.FILES['myfile'].name:
        filename = FileSystemStorage().save(survey_data_path, request.FILES['myfile'])
        # Rename headers
        missing_cols = rename_headers(survey_data_path)
        if missing_cols:
            return render(request, 'procedure/oops.html', {'error_list': ['Warning: the following columns could not be resolved from your survey file. These columns will be filled with blanks if you choose to continue:']+missing_cols, 'continue':True, 'continue_to':'step3'})
        else:
            return HttpResponseRedirect(reverse('step3'))
    return render(request, 'procedure/step2.html', {})

def step3(request):
    # Save Options
    if request.method == 'POST' and 'save' in request.POST:
        form = RunParametersForm(request.POST, request.FILES)
        if form.is_valid():
            params = form.save(commit=False)
            params.run_date = timezone.now()
            # If commutes uploaded, over-ride calc_commutes
            if params.commutes_reference.name:
                params.calc_commutes = False
            if not params.commutes_reference.name and params.calc_commutes == False:
                params.commute_factor = 0
            params.save()

            # Write parameters to csv
            params_fields = params._meta.get_fields()
            field_list = [field.name for field in params_fields]
            value_list = [str(getattr(params, field.name)) for field in params_fields]
            with open(params_path,'w', newline='') as f:
                w = csv.writer(f)
                w.writerow(field_list)
                w.writerow(value_list)

            # place Output_Commute_Reference.csv
            if params.commutes_reference.name and (os.path.normpath(params.commutes_reference.path) != os.path.normpath(commute_path)):
                if os.path.exists(commute_path):
                    os.remove(commute_path)
                os.rename(params.commutes_reference.path, commute_path)
            return HttpResponseRedirect(reverse('run'))
    else:
        form = RunParametersForm()
    return render(request, 'procedure/step3.html', {'form': form})

def run(request):
    # TODO: fails for certain encoding (like ANSI with certain characters)
    n_acms = sum(1 for line in open(survey_data_path)) - 1

    wb = load_workbook(school_data_path)
    sheet = wb.worksheets[0]
    n_schools = sheet.max_row

    params = RunParameters.objects.last()
    if params.calc_commutes == True:
        and_text = ' and calculating commutes'
        and_cost_text = ' and cost HQ '
        and_cost = f'${round(n_acms * n_schools * 0.005, 2)}'
    else:
        and_text, and_cost_text, and_cost = '', '', ''

    run_time = params.number_iterations/30 # filler: assumed 30 iterations per second
    if params.calc_commutes == True:
        run_time += n_acms*n_schools*0.5 # 0.5 seconds per API request
    run_time_mins = int(round(run_time/60))

    if request.method == 'POST' and 'run' in request.POST:
        return HttpResponseRedirect(reverse('dash'))

    return render(request, 'procedure/run.html', {'n_acms': n_acms, 'n_schools': n_schools, 'run_time_mins':run_time_mins, 'and_text':and_text, 'and_cost_text':and_cost_text, 'and_cost':and_cost})

def wait(request):
    #TODO: display progress
    return render(request, 'procedure/wait.html', {})

def dash(request):
    # download placements
    if request.method == 'POST' and 'download_placements' in request.POST:
        if os.path.exists(plcmt_path):
            with open(plcmt_path, 'rb') as fh:
                response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
                response['Content-Disposition'] = 'inline; filename=' + os.path.basename(plcmt_path)
                return response

    # download commutes
    if request.method == 'POST' and 'download_commutes' in request.POST:
        if os.path.exists(commute_path):
            with open(commute_path, 'rb') as fh:
                response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
                response['Content-Disposition'] = 'inline; filename=' + os.path.basename(commute_path)
                return response

    # Run Placement Process
    start_time = datetime.datetime.now()
    params = RunParameters.objects.last()
    # acm_df cleaned on upload
    # commutes in python:
    if params.calc_commutes == True:
        api = open('gdm_api_key.txt').readline()
        acm_df, school_df = clean_commute_inputs(survey_data_path, school_data_path, params.commute_date.strftime("%Y-%m-%d"))
        try:
            commute_procedure(acm_df, school_df, api, commute_path)
        except Exception as e:
            return render(request, 'procedure/oops.html', {'error_list': [str(e)], 'continue':False})

    # algorithm in R:
    os.system(f'Rscript --no-restore --no-save launch_alg.R > {error_path} 2>&1')
    run_time=round((datetime.datetime.now()-start_time).seconds/60, 1)
    # check for errors
    error_list = []
    with open(error_path) as error_text:
        for line in error_text:
            if line not in ['[[1]]\n', '[1] TRUE\n']:
                error_list.append(line)
    if 'execution halted' in str(error_list).lower():
        return render(request, 'procedure/oops.html', {'error_list': error_list, 'continue':False, 'commute_ref_present':os.path.exists(commute_path)})
    else:
        return render(request, 'procedure/dash.html', {'commute_ref_present':os.path.exists(commute_path), 'run_time':run_time})

def oops(request):
    # download commutes
    if request.method == 'POST' and 'download_commutes' in request.POST:
        if os.path.exists(commute_path):
            with open(commute_path, 'rb') as fh:
                response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
                response['Content-Disposition'] = 'inline; filename=' + os.path.basename(commute_path)
                return response
    return render(request, 'procedure/oops.html', {})
